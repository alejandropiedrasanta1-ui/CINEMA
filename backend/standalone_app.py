"""
Cinema Productions - App Local Auto-contenida
==============================================
Ejecutar: python app.py  (o doble clic en start.bat)

Modos de base de datos:
  MONGO_URL=embedded          → Almacenamiento local en cinema_data.json (predeterminado)
  MONGO_URL=mongodb://...     → MongoDB externo (local o Atlas)
"""
from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, File
from fastapi.responses import JSONResponse, Response, FileResponse
from fastapi.staticfiles import StaticFiles
from starlette.middleware.cors import CORSMiddleware
from contextlib import asynccontextmanager
from dotenv import load_dotenv
import csv
import io
import os
import re
import json
import logging
import base64
import uuid
import webbrowser
import threading
import time
import asyncio
from pathlib import Path
from pydantic import BaseModel, Field
from typing import Optional
from datetime import datetime, timezone, timedelta
from bson import ObjectId

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

DB_NAME = os.environ.get('DB_NAME', 'cinema_productions')
MONGO_URL = os.environ.get('MONGO_URL', 'embedded')
DATA_FILE = ROOT_DIR / 'cinema_data.json'
CUSTOM_DB_FILE = ROOT_DIR / '.db_override'
BACKUP_DIR = ROOT_DIR / 'backups'
BACKUP_DIR.mkdir(exist_ok=True)
BACKUP_COLLECTIONS = ['reservations', 'socios', 'app_settings']

# ── Determine effective MongoDB URL (override file takes priority) ────────────
_override_url = CUSTOM_DB_FILE.read_text().strip() if CUSTOM_DB_FILE.exists() else None
_effective_mongo_url = _override_url or MONGO_URL

_using_embedded = _effective_mongo_url.strip().lower() in ('embedded', '', 'local')

logging.basicConfig(level=logging.WARNING)
logger = logging.getLogger(__name__)

if _using_embedded:
    from mongomock_motor import AsyncMongoMockClient
    _mongo_client = AsyncMongoMockClient()
    logger.warning("Usando base de datos embebida (cinema_data.json)")
else:
    try:
        from motor.motor_asyncio import AsyncIOMotorClient as _MotorClient
        _mongo_client = _MotorClient(_effective_mongo_url, serverSelectionTimeoutMS=8000)
        logger.warning(f"Motor inicializado para: {_effective_mongo_url[:50]}...")
    except Exception as _e:
        logger.error(f"Error al inicializar motor: {_e} — usando modo embebido")
        from mongomock_motor import AsyncMongoMockClient
        _mongo_client = AsyncMongoMockClient()
        _using_embedded = True

db = _mongo_client[DB_NAME]
client = _mongo_client  # alias


# ─── Embedded DB persistence ─────────────────────────────

def _serialize_doc(doc: dict) -> dict:
    """Convert MongoDB doc to JSON-serializable dict."""
    d = {}
    for k, v in doc.items():
        if k == '_id':
            d['__id'] = str(v)
        elif isinstance(v, ObjectId):
            d[k] = str(v)
        elif isinstance(v, list):
            d[k] = [_serialize_doc(i) if isinstance(i, dict) else i for i in v]
        else:
            d[k] = v
    return d


def _deserialize_doc(doc: dict) -> dict:
    """Restore JSON doc to MongoDB-compatible dict with ObjectId _id."""
    d = {}
    for k, v in doc.items():
        if k == '__id':
            try:
                d['_id'] = ObjectId(v)
            except Exception:
                d['_id'] = v
        elif isinstance(v, list):
            d[k] = [_deserialize_doc(i) if isinstance(i, dict) else i for i in v]
        else:
            d[k] = v
    return d


async def _load_embedded_data():
    if not DATA_FILE.exists():
        logger.info("No saved data found — starting fresh.")
        return
    try:
        data = json.loads(DATA_FILE.read_text(encoding='utf-8'))
        for col_name, docs in data.items():
            if not docs:
                continue
            col = db[col_name]
            restored = [_deserialize_doc(d) for d in docs]
            if restored:
                await col.insert_many(restored)
        logger.info(f"Loaded {sum(len(v) for v in data.values())} docs from {DATA_FILE.name}")
    except Exception as e:
        logger.error(f"Failed to load saved data: {e}")


async def _save_embedded_data():
    try:
        data = {}
        for col_name in ['reservations', 'socios', 'app_settings']:
            docs = await db[col_name].find({}).to_list(100000)
            data[col_name] = [_serialize_doc(d) for d in docs]
        DATA_FILE.write_text(
            json.dumps(data, ensure_ascii=False, indent=2, default=str),
            encoding='utf-8'
        )
        logger.info(f"Data saved → {DATA_FILE.name}")
    except Exception as e:
        logger.error(f"Failed to save data: {e}")


async def _auto_save_loop():
    """Auto-save every 60 seconds."""
    while True:
        await asyncio.sleep(60)
        await _save_embedded_data()


# ─── Lifespan ─────────────────────────────────────────────

@asynccontextmanager
async def lifespan(app_instance: FastAPI):
    global db, client, _using_embedded
    _task = None

    # ── Verify MongoDB connection or fall back to embedded ──────────────────
    if not _using_embedded:
        try:
            await asyncio.wait_for(db.command("ping"), timeout=9.0)
            logger.warning("MongoDB Atlas: conexión verificada OK")
        except Exception as _conn_err:
            logger.error(
                f"MongoDB no accesible ({_conn_err}). "
                "Revisa tu conexión a internet o el archivo .env. "
                "Activando modo embebido temporal."
            )
            from mongomock_motor import AsyncMongoMockClient
            _fb = AsyncMongoMockClient()
            db = _fb[DB_NAME]
            client = _fb
            _using_embedded = True

    if _using_embedded:
        await _load_embedded_data()
        _task = asyncio.create_task(_auto_save_loop())
    yield
    if _using_embedded:
        if _task:
            _task.cancel()
        await _save_embedded_data()
    _mongo_client.close()


app = FastAPI(title="Cinema Productions", lifespan=lifespan)
api_router = APIRouter(prefix="/api")


# ─── Helpers ──────────────────────────────────────────────

def doc_to_dict(doc: dict) -> dict:
    if doc is None:
        return {}
    d = {k: v for k, v in doc.items() if k != '_id'}
    if '_id' in doc:
        d['id'] = str(doc['_id'])
    return d


# ─── Models ───────────────────────────────────────────────

class ReservationCreate(BaseModel):
    client_name: str
    client_phone: Optional[str] = None
    client_email: Optional[str] = None
    event_type: str
    event_date: str
    event_time: Optional[str] = None
    venue: Optional[str] = None
    guests_count: Optional[int] = None
    total_amount: float
    advance_paid: float = 0.0
    status: str = "Pendiente"
    notes: Optional[str] = None


class ReservationUpdate(BaseModel):
    client_name: Optional[str] = None
    client_phone: Optional[str] = None
    client_email: Optional[str] = None
    event_type: Optional[str] = None
    event_date: Optional[str] = None
    event_time: Optional[str] = None
    venue: Optional[str] = None
    guests_count: Optional[int] = None
    total_amount: Optional[float] = None
    advance_paid: Optional[float] = None
    status: Optional[str] = None
    notes: Optional[str] = None
    locations: Optional[list] = None
    assigned_partners: Optional[list] = None


class SocioCreate(BaseModel):
    name: str
    role: str = "Fotógrafo"
    phone: Optional[str] = None
    email: Optional[str] = None
    notes: Optional[str] = None
    rate_per_event: Optional[float] = None


class SocioUpdate(BaseModel):
    name: Optional[str] = None
    role: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    notes: Optional[str] = None
    rate_per_event: Optional[float] = None


class NotificationSettingsModel(BaseModel):
    reminders_enabled: bool = False
    reminder_days: Optional[int] = 3
    reminder_periods: Optional[list] = None
    reminder_time: Optional[str] = "09:00"
    reminder_hours_before: Optional[int] = 0
    admin_email: Optional[str] = None
    admin_whatsapp: Optional[str] = None
    notification_channel: str = "email"
    resend_api_key: Optional[str] = None
    resend_sender_name: Optional[str] = None
    resend_cc: Optional[str] = None
    resend_subject_template: Optional[str] = None
    notify_client: Optional[bool] = False
    telegram_enabled: Optional[bool] = False
    telegram_bot_token: Optional[str] = None
    telegram_chat_id: Optional[str] = None
    ntfy_enabled: Optional[bool] = False
    ntfy_topic: Optional[str] = None
    company_name: Optional[str] = None
    company_address: Optional[str] = None
    company_phone: Optional[str] = None
    company_email: Optional[str] = None
    company_logo: Optional[str] = None
    currency: Optional[str] = "Q"
    language: Optional[str] = "es"
    timezone: Optional[str] = None


class DBConnectRequest(BaseModel):
    url: str


# ─── Routes ───────────────────────────────────────────────

@api_router.get("/")
async def root():
    mode = "embedded" if _using_embedded else "mongodb"
    return {"message": "Cinema Productions API", "db_mode": mode}


@api_router.get("/stats")
async def get_stats():
    total = await db.reservations.count_documents({})
    today_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    all_res = await db.reservations.find(
        {}, {"event_date": 1, "status": 1, "total_amount": 1, "advance_paid": 1, "_id": 0}
    ).to_list(10000)
    upcoming = sum(1 for d in all_res
                   if d.get("event_date", "") >= today_str
                   and d.get("status") not in ("Cancelado", "Completado"))
    total_pending = sum(
        max(0, (d.get("total_amount") or 0) - (d.get("advance_paid") or 0))
        for d in all_res if d.get("status") not in ("Cancelado",)
    )
    confirmed = sum(1 for d in all_res if d.get("status") == "Confirmado")
    return {
        "total_reservations": total,
        "upcoming_events": upcoming,
        "pending_payment": round(total_pending, 2),
        "confirmed": confirmed
    }


@api_router.get("/reservations")
async def list_reservations():
    docs = await db.reservations.find({}, {"receipt_images.data": 0}).to_list(1000)
    return [doc_to_dict(d) for d in docs]


@api_router.post("/reservations", status_code=201)
async def create_reservation(reservation: ReservationCreate):
    doc = reservation.model_dump()
    doc["receipt_images"] = []
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.reservations.insert_one(doc)
    doc["id"] = str(result.inserted_id)
    doc.pop("_id", None)
    if _using_embedded:
        asyncio.create_task(_save_embedded_data())
    return doc


@api_router.get("/reservations/{reservation_id}")
async def get_reservation(reservation_id: str):
    try:
        oid = ObjectId(reservation_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    doc = await db.reservations.find_one({"_id": oid})
    if not doc:
        raise HTTPException(status_code=404, detail="Reserva no encontrada")
    return doc_to_dict(doc)


@api_router.put("/reservations/{reservation_id}")
async def update_reservation(reservation_id: str, reservation: ReservationUpdate):
    try:
        oid = ObjectId(reservation_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    update_data = {k: v for k, v in reservation.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="Sin datos para actualizar")
    result = await db.reservations.update_one({"_id": oid}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Reserva no encontrada")
    doc = await db.reservations.find_one({"_id": oid})
    if _using_embedded:
        asyncio.create_task(_save_embedded_data())
    return doc_to_dict(doc)


@api_router.delete("/reservations/{reservation_id}")
async def delete_reservation(reservation_id: str):
    try:
        oid = ObjectId(reservation_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    result = await db.reservations.delete_one({"_id": oid})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Reserva no encontrada")
    if _using_embedded:
        asyncio.create_task(_save_embedded_data())
    return {"message": "Reserva eliminada"}


@api_router.post("/reservations/{reservation_id}/receipts")
async def upload_receipt(reservation_id: str, file: UploadFile = File(...)):
    try:
        oid = ObjectId(reservation_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Archivo muy grande (máx 10MB)")
    b64 = base64.b64encode(content).decode("utf-8")
    receipt = {
        "id": str(uuid.uuid4()),
        "filename": file.filename,
        "content_type": file.content_type,
        "data": b64,
        "uploaded_at": datetime.now(timezone.utc).isoformat()
    }
    result = await db.reservations.update_one({"_id": oid}, {"$push": {"receipt_images": receipt}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Reserva no encontrada")
    if _using_embedded:
        asyncio.create_task(_save_embedded_data())
    return {k: v for k, v in receipt.items() if k != "data"}


@api_router.delete("/reservations/{reservation_id}/receipts/{receipt_id}")
async def delete_receipt(reservation_id: str, receipt_id: str):
    try:
        oid = ObjectId(reservation_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    result = await db.reservations.update_one(
        {"_id": oid}, {"$pull": {"receipt_images": {"id": receipt_id}}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Reserva no encontrada")
    if _using_embedded:
        asyncio.create_task(_save_embedded_data())
    return {"message": "Comprobante eliminado"}


@api_router.get("/export/reservations")
async def export_reservations(format: str = "csv"):
    docs = await db.reservations.find({}, {"receipt_images": 0}).to_list(10000)
    data = [doc_to_dict(d) for d in docs]
    if format == "json":
        return JSONResponse(content=data,
                            headers={"Content-Disposition": "attachment; filename=reservaciones.json"})
    if not data:
        return Response("", media_type="text/csv",
                        headers={"Content-Disposition": "attachment; filename=reservaciones.csv"})
    output = io.StringIO()
    fields = ["id", "client_name", "client_phone", "client_email", "event_type", "event_date",
              "event_time", "venue", "guests_count", "total_amount", "advance_paid", "status", "notes", "created_at"]
    writer = csv.DictWriter(output, fieldnames=fields, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(data)
    return Response(output.getvalue(), media_type="text/csv; charset=utf-8",
                    headers={"Content-Disposition": "attachment; filename=reservaciones.csv"})


@api_router.get("/calendar")
async def get_calendar_events():
    docs = await db.reservations.find(
        {"status": {"$nin": ["Cancelado"]}},
        {"client_name": 1, "event_date": 1, "event_type": 1, "status": 1, "_id": 1}
    ).to_list(1000)
    return [doc_to_dict(d) for d in docs]


# ─── Socios ───────────────────────────────────────────────

@api_router.get("/socios")
async def list_socios():
    docs = await db.socios.find({}, {"photo": 0, "photo_content_type": 0}).to_list(1000)
    return [doc_to_dict(d) for d in docs]


@api_router.post("/socios", status_code=201)
async def create_socio(socio: SocioCreate):
    doc = socio.model_dump()
    doc.update({"photo": None, "photo_content_type": None,
                "created_at": datetime.now(timezone.utc).isoformat()})
    result = await db.socios.insert_one(doc)
    doc["id"] = str(result.inserted_id)
    doc.pop("_id", None)
    if _using_embedded:
        asyncio.create_task(_save_embedded_data())
    return doc


@api_router.get("/socios/{socio_id}")
async def get_socio(socio_id: str):
    try:
        oid = ObjectId(socio_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    doc = await db.socios.find_one({"_id": oid})
    if not doc:
        raise HTTPException(status_code=404, detail="Socio no encontrado")
    return doc_to_dict(doc)


@api_router.put("/socios/{socio_id}")
async def update_socio(socio_id: str, socio: SocioUpdate):
    try:
        oid = ObjectId(socio_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    update_data = {k: v for k, v in socio.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="Sin datos para actualizar")
    result = await db.socios.update_one({"_id": oid}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Socio no encontrado")
    doc = await db.socios.find_one({"_id": oid})
    if _using_embedded:
        asyncio.create_task(_save_embedded_data())
    return doc_to_dict(doc)


@api_router.delete("/socios/{socio_id}")
async def delete_socio(socio_id: str):
    try:
        oid = ObjectId(socio_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    result = await db.socios.delete_one({"_id": oid})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Socio no encontrado")
    if _using_embedded:
        asyncio.create_task(_save_embedded_data())
    return {"message": "Socio eliminado"}


@api_router.post("/socios/{socio_id}/photo")
async def upload_socio_photo(socio_id: str, file: UploadFile = File(...)):
    try:
        oid = ObjectId(socio_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Archivo muy grande (máx 5MB)")
    b64 = base64.b64encode(content).decode("utf-8")
    result = await db.socios.update_one(
        {"_id": oid}, {"$set": {"photo": b64, "photo_content_type": file.content_type}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Socio no encontrado")
    if _using_embedded:
        asyncio.create_task(_save_embedded_data())
    return {"message": "Foto actualizada"}


@api_router.delete("/socios/{socio_id}/photo")
async def delete_socio_photo(socio_id: str):
    try:
        oid = ObjectId(socio_id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    await db.socios.update_one({"_id": oid}, {"$set": {"photo": None, "photo_content_type": None}})
    return {"message": "Foto eliminada"}


@api_router.get("/financials")
async def get_financials():
    docs = await db.reservations.find(
        {"status": {"$nin": ["Cancelado"]}},
        {"total_amount": 1, "advance_paid": 1, "assigned_partners": 1}
    ).to_list(10000)
    total_event = sum((d.get("total_amount") or 0) for d in docs)
    total_advance = sum((d.get("advance_paid") or 0) for d in docs)
    total_cost = total_paid = total_pending = 0
    for d in docs:
        for p in (d.get("assigned_partners") or []):
            amt = p.get("payment") or 0
            total_cost += amt
            if p.get("payment_status") == "Pagado":
                total_paid += amt
            else:
                total_pending += amt
    return {
        "total_event_amount": round(total_event, 2),
        "total_advance": round(total_advance, 2),
        "total_partner_cost": round(total_cost, 2),
        "total_paid_to_partners": round(total_paid, 2),
        "total_pending_to_partners": round(total_pending, 2),
        "real_income": round(total_event - total_cost, 2),
    }


@api_router.get("/settings")
async def get_app_settings():
    doc = await db.app_settings.find_one({}, {"_id": 0})
    if not doc:
        return {}
    if doc.get("resend_api_key"):
        key = doc["resend_api_key"]
        doc["resend_api_key"] = "re_" + "*" * 20 + key[-4:] if len(key) > 4 else "****"
        doc["has_resend_key"] = True
    else:
        doc["has_resend_key"] = False
    return doc


@api_router.put("/settings")
async def update_app_settings(settings: NotificationSettingsModel):
    update_doc = {k: v for k, v in settings.model_dump().items() if v is not None}
    # Don't overwrite the real key if frontend sends a masked placeholder
    key = update_doc.get("resend_api_key") or ""
    if "****" in key or key.startswith("re_" + "*"):
        update_doc.pop("resend_api_key", None)
    telegram_token = update_doc.get("telegram_bot_token") or ""
    if "•" in telegram_token or "****" in telegram_token:
        update_doc.pop("telegram_bot_token", None)
    existing = await db.app_settings.find_one({}, {"_id": 0})
    if existing:
        await db.app_settings.update_one({}, {"$set": update_doc})
    else:
        await db.app_settings.insert_one(update_doc)
    if _using_embedded:
        asyncio.create_task(_save_embedded_data())
    saved = await db.app_settings.find_one({}, {"_id": 0})
    if saved and saved.get("resend_api_key"):
        key = saved["resend_api_key"]
        saved["resend_api_key"] = "re_" + "*" * 20 + key[-4:] if len(key) > 4 else "****"
        saved["has_resend_key"] = True
    elif saved:
        saved["has_resend_key"] = False
    return saved or {}


@api_router.get("/settings/database")
async def get_db_stats():
    if _using_embedded:
        counts = {}
        total_docs = 0
        for col in ['reservations', 'socios', 'app_settings']:
            n = await db[col].count_documents({})
            counts[col] = n
            total_docs += n
        data_file_size = DATA_FILE.stat().st_size if DATA_FILE.exists() else 0

        def fmt(b):
            if b < 1024:
                return f"{b} B"
            elif b < 1024 ** 2:
                return f"{b / 1024:.1f} KB"
            return f"{b / (1024 ** 2):.2f} MB"

        return {
            "db_name": DB_NAME + " (embebida)",
            "collections": len(counts),
            "objects": total_docs,
            "data_size": fmt(data_file_size),
            "storage_size": fmt(data_file_size),
            "index_size": "0 B",
            "total_size": fmt(data_file_size),
            "current_url": "embedded (cinema_data.json)",
            "is_custom": False,
        }
    try:
        raw = await db.command("dbstats")
        storage_bytes = raw.get("storageSize", 0)
        data_bytes = raw.get("dataSize", 0)
        index_bytes = raw.get("indexSize", 0)

        def fmt(b):
            if b < 1024:
                return f"{b} B"
            elif b < 1024 ** 2:
                return f"{b / 1024:.1f} KB"
            return f"{b / (1024 ** 2):.2f} MB"

        is_custom = CUSTOM_DB_FILE.exists()
        current_url = CUSTOM_DB_FILE.read_text().strip() if is_custom else MONGO_URL
        display_url = current_url
        if "@" in current_url:
            proto_end = current_url.find("://") + 3
            at_pos = current_url.rfind("@")
            display_url = current_url[:proto_end] + "***:***@" + current_url[at_pos + 1:]
        return {
            "db_name": DB_NAME,
            "collections": raw.get("collections", 0),
            "objects": raw.get("objects", 0),
            "data_size": fmt(data_bytes),
            "storage_size": fmt(storage_bytes),
            "index_size": fmt(index_bytes),
            "total_size": fmt(storage_bytes + index_bytes),
            "current_url": display_url,
            "is_custom": is_custom,
        }
    except Exception as e:
        # Atlas not reachable — return safe fallback so UI doesn't crash
        is_custom = CUSTOM_DB_FILE.exists()
        current_url = CUSTOM_DB_FILE.read_text().strip() if is_custom else _effective_mongo_url
        display_url = current_url
        if "@" in current_url:
            proto_end = current_url.find("://") + 3
            at_pos = current_url.rfind("@")
            display_url = current_url[:proto_end] + "***:***@" + current_url[at_pos + 1:]
        return {
            "db_name": DB_NAME,
            "collections": 0,
            "objects": 0,
            "data_size": "—",
            "storage_size": "—",
            "index_size": "—",
            "total_size": "—",
            "current_url": display_url,
            "is_custom": is_custom,
            "connection_error": str(e),
        }


@api_router.post("/settings/database/test")
async def test_db_connection(req: DBConnectRequest):
    if req.url.lower() == "embedded":
        return {"success": True, "message": "Modo embebido seleccionado"}
    try:
        from motor.motor_asyncio import AsyncIOMotorClient
        test_client = AsyncIOMotorClient(req.url, serverSelectionTimeoutMS=5000)
        await test_client[DB_NAME].command("ping")
        test_client.close()
        return {"success": True, "message": "Conexión exitosa"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo conectar: {e}")


@api_router.post("/settings/database/connect")
async def switch_database(req: DBConnectRequest):
    global db, client, _using_embedded
    url = req.url.strip()

    # Switch to embedded mode
    if url.lower() in ('embedded', '', 'local'):
        from mongomock_motor import AsyncMongoMockClient
        new_client = AsyncMongoMockClient()
        db = new_client[DB_NAME]
        client = new_client
        CUSTOM_DB_FILE.unlink(missing_ok=True)
        _using_embedded = True
        return {"success": True, "message": "Modo embebido activado. Los datos se guardan localmente."}

    # Switch to MongoDB URL
    try:
        from motor.motor_asyncio import AsyncIOMotorClient as _MotorClient
        new_client = _MotorClient(url, serverSelectionTimeoutMS=8000)
        await new_client[DB_NAME].command("ping")
        db = new_client[DB_NAME]
        client = new_client
        CUSTOM_DB_FILE.write_text(url)
        _using_embedded = False
        return {"success": True, "message": "Base de datos conectada correctamente"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error al conectar: {e}")


@api_router.post("/settings/database/reset")
async def reset_database():
    global db, client, _using_embedded
    if CUSTOM_DB_FILE.exists():
        CUSTOM_DB_FILE.unlink()
    # Fall back to default MONGO_URL from .env
    default_url = MONGO_URL.strip()
    if default_url.lower() in ('embedded', '', 'local'):
        from mongomock_motor import AsyncMongoMockClient
        new_client = AsyncMongoMockClient()
        _using_embedded = True
    else:
        from motor.motor_asyncio import AsyncIOMotorClient as _MotorClient
        new_client = _MotorClient(default_url)
        _using_embedded = False
    db = new_client[DB_NAME]
    client = new_client
    return {"success": True, "message": "Restaurado a la base de datos predeterminada."}


@api_router.delete("/data/clear-all")
async def clear_all_data_endpoint():
    try:
        await _create_backup(label="auto_pre_delete")
    except Exception:
        pass
    res_result = await db.reservations.delete_many({})
    soc_result = await db.socios.delete_many({})
    await db.app_settings.delete_many({})
    if _using_embedded:
        asyncio.create_task(_save_embedded_data())
    return {
        "ok": True,
        "deleted_reservations": res_result.deleted_count,
        "deleted_socios": soc_result.deleted_count,
        "auto_backup_created": True,
    }


@api_router.post("/data/cleanup")
async def cleanup_data(action: str = "cancelled", months_old: int = 6):
    if action == "preview":
        cancelled_count = await db.reservations.count_documents({"status": "Cancelado"})
        cutoff = datetime.now(timezone.utc) - timedelta(days=months_old * 30)
        cutoff_str = cutoff.strftime("%Y-%m-%d")
        docs = await db.reservations.find({"status": "Completado"}).to_list(100000)
        old_completed = sum(1 for d in docs if d.get("event_date", "") < cutoff_str)
        return {
            "ok": True,
            "cancelled_count": cancelled_count,
            "old_completed_count": old_completed,
            "months_threshold": months_old,
        }

    try:
        await _create_backup(label="auto_pre_cleanup")
    except Exception:
        pass

    if action == "cancelled":
        result = await db.reservations.delete_many({"status": "Cancelado"})
        if _using_embedded:
            asyncio.create_task(_save_embedded_data())
        return {"ok": True, "deleted": result.deleted_count, "message": f"{result.deleted_count} reservas canceladas eliminadas"}

    if action == "old_completed":
        cutoff = datetime.now(timezone.utc) - timedelta(days=months_old * 30)
        cutoff_str = cutoff.strftime("%Y-%m-%d")
        docs = await db.reservations.find({"status": "Completado"}).to_list(100000)
        ids_to_delete = [d["_id"] for d in docs if d.get("event_date", "") < cutoff_str]
        if ids_to_delete:
            result = await db.reservations.delete_many({"_id": {"$in": ids_to_delete}})
            if _using_embedded:
                asyncio.create_task(_save_embedded_data())
            return {"ok": True, "deleted": result.deleted_count, "message": f"{result.deleted_count} reservas completadas antiguas eliminadas"}
        return {"ok": True, "deleted": 0, "message": "No hay reservas completadas antiguas para eliminar"}

    return {"ok": False, "message": "Acción no reconocida"}


# ─── Backup helper ─────────────────────────────────────────

async def _create_backup(label: str = "manual") -> dict:
    backup_data: dict = {"_meta": {"created_at": datetime.now(timezone.utc).isoformat(), "label": label}}
    for cname in BACKUP_COLLECTIONS:
        docs = await db[cname].find({}).to_list(100000)
        backup_data[cname] = [doc_to_dict(d) for d in docs]
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"backup_{label}_{ts}.json"
    filepath = BACKUP_DIR / filename
    filepath.write_text(json.dumps(backup_data, ensure_ascii=False, indent=2), encoding="utf-8")
    # Keep only last 15 backups
    existing = sorted(BACKUP_DIR.glob("backup_*.json"), key=lambda f: f.stat().st_mtime)
    for old in existing[:-15]:
        old.unlink(missing_ok=True)
    total_docs = sum(len(v) for k, v in backup_data.items() if k != "_meta")
    return {"filename": filename, "docs": total_docs}


# ─── Backup endpoints ──────────────────────────────────────

@api_router.get("/backup/download")
async def download_full_backup():
    backup_data: dict = {"_meta": {"created_at": datetime.now(timezone.utc).isoformat(), "app": "Cinema Productions"}}
    for cname in BACKUP_COLLECTIONS:
        docs = await db[cname].find({}).to_list(100000)
        backup_data[cname] = [doc_to_dict(d) for d in docs]
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"cinema_backup_{ts}.json"
    content = json.dumps(backup_data, ensure_ascii=False, indent=2)
    return Response(
        content=content.encode("utf-8"),
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@api_router.post("/backup/create")
async def create_server_backup():
    try:
        result = await _create_backup(label="manual")
        return {"success": True, **result, "message": f"Respaldo creado: {result['docs']} documentos"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al crear respaldo: {e}")


@api_router.get("/backup/history")
async def list_backups():
    files = sorted(BACKUP_DIR.glob("backup_*.json"), key=lambda f: f.stat().st_mtime, reverse=True)
    result = []
    for f in files:
        stat = f.stat()
        size_kb = stat.st_size / 1024
        size_str = f"{size_kb:.0f} KB" if size_kb < 1024 else f"{size_kb / 1024:.1f} MB"
        label = "auto" if "auto_" in f.name else "manual"
        result.append({
            "filename": f.name,
            "size": size_str,
            "label": label,
            "created_at": datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat(),
        })
    return result


@api_router.get("/backup/{filename}/download")
async def download_backup_file(filename: str):
    if not filename.endswith(".json") or "/" in filename or ".." in filename:
        raise HTTPException(status_code=400, detail="Nombre de archivo inválido")
    filepath = BACKUP_DIR / filename
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Respaldo no encontrado")
    return Response(
        content=filepath.read_bytes(),
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@api_router.delete("/backup/{filename}")
async def delete_backup_file(filename: str):
    if not filename.endswith(".json") or "/" in filename or ".." in filename:
        raise HTTPException(status_code=400, detail="Nombre de archivo inválido")
    filepath = BACKUP_DIR / filename
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Respaldo no encontrado")
    filepath.unlink()
    return {"success": True, "message": "Respaldo eliminado"}


@api_router.post("/backup/restore")
async def restore_backup(file: UploadFile = File(...)):
    if not file.filename.endswith(".json"):
        raise HTTPException(status_code=400, detail="El archivo debe ser .json")
    try:
        content = await file.read()
        backup_data = json.loads(content)
    except Exception:
        raise HTTPException(status_code=400, detail="Archivo JSON inválido o corrupto")
    try:
        await _create_backup(label="auto_pre_restore")
    except Exception:
        pass
    restored: dict = {}
    errors: list = []
    for cname in BACKUP_COLLECTIONS:
        docs = backup_data.get(cname)
        if docs is None or not isinstance(docs, list):
            continue
        try:
            clean_docs = []
            for d in docs:
                d.pop("id", None)
                d.pop("_id", None)
                clean_docs.append(d)
            await db[cname].delete_many({})
            if clean_docs:
                await db[cname].insert_many(clean_docs)
            restored[cname] = len(clean_docs)
        except Exception as e:
            errors.append(f"{cname}: {e}")
    if errors:
        raise HTTPException(status_code=500, detail="Errores al restaurar: " + " | ".join(errors))
    if _using_embedded:
        asyncio.create_task(_save_embedded_data())
    total = sum(restored.values())
    return {"success": True, "restored": restored, "total": total,
            "message": f"Restaurado correctamente: {total} documentos en {len(restored)} colecciones"}


# ─── Import / Export adicionales ──────────────────────────

@api_router.post("/import/reservations")
async def import_reservations_csv(file: UploadFile = File(...)):
    content = await file.read()
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")
    reader = csv.DictReader(io.StringIO(text))
    FIELD_MAP = {
        "client_name":  ["client_name", "nombre", "cliente", "name"],
        "client_phone": ["client_phone", "telefono", "teléfono", "phone"],
        "client_email": ["client_email", "email", "correo"],
        "event_type":   ["event_type", "tipo_evento", "tipo", "type"],
        "event_date":   ["event_date", "fecha", "date", "fecha_evento"],
        "event_time":   ["event_time", "hora", "time"],
        "venue":        ["venue", "lugar", "location", "ubicacion"],
        "guests_count": ["guests_count", "invitados", "guests"],
        "total_amount": ["total_amount", "total", "monto_total"],
        "advance_paid": ["advance_paid", "anticipo", "advance", "pago_anticipado"],
        "status":       ["status", "estado"],
        "notes":        ["notes", "notas", "note"],
    }

    def find_col(headers, candidates):
        headers_lower = {h.lower().strip(): h for h in headers}
        for c in candidates:
            if c.lower() in headers_lower:
                return headers_lower[c.lower()]
        return None

    imported = 0
    errors = []
    now_str = datetime.now(timezone.utc).isoformat()

    for i, row in enumerate(reader, start=2):
        try:
            headers = list(row.keys())
            def get(field):
                col = find_col(headers, FIELD_MAP.get(field, [field]))
                return row.get(col, "").strip() if col else ""
            client_name = get("client_name")
            if not client_name:
                errors.append(f"Fila {i}: nombre vacío — omitida")
                continue

            def to_float(s):
                s = re.sub(r"[^\d.]", "", s)
                return float(s) if s else 0.0

            doc = {
                "client_name":   client_name,
                "client_phone":  get("client_phone") or None,
                "client_email":  get("client_email") or None,
                "event_type":    get("event_type") or "Otro",
                "event_date":    get("event_date") or "",
                "event_time":    get("event_time") or None,
                "venue":         get("venue") or None,
                "guests_count":  int(get("guests_count")) if get("guests_count").isdigit() else None,
                "total_amount":  to_float(get("total_amount") or "0"),
                "advance_paid":  to_float(get("advance_paid") or "0"),
                "status":        get("status") or "Pendiente",
                "notes":         get("notes") or None,
                "receipts":      [],
                "locations":     [],
                "assigned_partners": [],
                "created_at":    now_str,
                "imported_from_csv": True,
            }
            await db.reservations.insert_one(doc)
            imported += 1
        except Exception as e:
            errors.append(f"Fila {i}: {str(e)}")

    if _using_embedded:
        asyncio.create_task(_save_embedded_data())
    return {
        "ok": True,
        "imported": imported,
        "errors": errors[:10],
        "message": f"{imported} reservas importadas correctamente" + (f" ({len(errors)} errores)" if errors else ""),
    }


@api_router.get("/export/reservations/xlsx")
async def export_reservations_xlsx():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        import io as _io

        docs = await db.reservations.find({}, {"receipt_images": 0, "assigned_partners": 0}).to_list(100000)
        data = [doc_to_dict(d) for d in docs]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reservas"
        headers = ["Nombre", "Teléfono", "Email", "Tipo Evento", "Fecha", "Hora",
                   "Lugar", "Invitados", "Total", "Anticipo", "Saldo", "Estado", "Notas", "Creado"]
        keys = ["client_name", "client_phone", "client_email", "event_type", "event_date",
                "event_time", "venue", "guests_count", "total_amount", "advance_paid",
                None, "status", "notes", "created_at"]
        header_fill = PatternFill("solid", fgColor="4F46E5")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        alt_fill = PatternFill("solid", fgColor="F8F7FF")
        for row_idx, doc in enumerate(data, start=2):
            fill = alt_fill if row_idx % 2 == 0 else None
            for col_idx, key in enumerate(keys, start=1):
                if key is None:
                    value = (doc.get("total_amount") or 0) - (doc.get("advance_paid") or 0)
                else:
                    value = doc.get(key, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if fill:
                    cell.fill = fill
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)
        buf = _io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="reservaciones_{ts}.xlsx"'},
        )
    except ImportError:
        return JSONResponse({"error": "openpyxl no instalado. Ejecuta: pip install openpyxl"}, status_code=500)
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)


@api_router.post("/reminders/send")
async def trigger_reminders_manual():
    return {"success": True, "events_found": 0, "sent": 0,
            "message": "Recordatorios por email disponibles en la versión web"}


# ─── App config ───────────────────────────────────────────

app.include_router(api_router)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ─── Serve React build (SPA) ─────────────────────────────

_LOCAL_INJECT = '<script>window.__API_BASE_URL__="http://localhost:8001";</script>'


def _inject_local_url(html: str) -> str:
    """Inject the local API URL so any build (old or new) works offline."""
    return html.replace("</head>", _LOCAL_INJECT + "</head>", 1)


BUILD_DIR = ROOT_DIR / "build"
if BUILD_DIR.exists():
    app.mount("/static", StaticFiles(directory=str(BUILD_DIR / "static")), name="static")

    @app.get("/favicon.ico")
    async def favicon():
        f = BUILD_DIR / "favicon.ico"
        return FileResponse(str(f)) if f.exists() else Response(status_code=204)

    @app.get("/manifest.json")
    async def manifest_json():
        f = BUILD_DIR / "manifest.json"
        return FileResponse(str(f)) if f.exists() else Response("{}", media_type="application/json")

    @app.get("/")
    async def serve_index():
        html_path = BUILD_DIR / "index.html"
        html = _inject_local_url(html_path.read_text(encoding="utf-8"))
        return Response(content=html, media_type="text/html; charset=utf-8")

    @app.get("/{path:path}")
    async def serve_spa(path: str):
        f = BUILD_DIR / path
        if f.exists() and f.is_file():
            return FileResponse(str(f))
        html_path = BUILD_DIR / "index.html"
        html = _inject_local_url(html_path.read_text(encoding="utf-8"))
        return Response(content=html, media_type="text/html; charset=utf-8")


# ─── Entry point ─────────────────────────────────────────

if __name__ == "__main__":
    import uvicorn

    db_label = "Embebida (cinema_data.json)" if _using_embedded else MONGO_URL[:40]

    def _open_browser():
        time.sleep(3)
        webbrowser.open("http://localhost:8001")

    threading.Thread(target=_open_browser, daemon=True).start()

    print("\n" + "=" * 54)
    print("  CINEMA PRODUCTIONS — Gestor de Reservas")
    print("=" * 54)
    print(f"  URL:  http://localhost:8001")
    print(f"  BD:   {db_label}")
    print(f"  Datos: {DATA_FILE.name if _using_embedded else 'MongoDB'}")
    print("  Para cerrar: Ctrl+C  o  cierra esta ventana")
    print("=" * 54 + "\n")

    uvicorn.run(app, host="0.0.0.0", port=8001, log_level="warning")
